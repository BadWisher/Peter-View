"""Генерация xlsx-отчёта по результатам проверки."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


SEVERITY_COLORS = {
    "error": "FFCCCC",
    "warning": "FFF3CD",
    "suggestion": "CCE5FF",
}

SEVERITY_LABELS = {
    "error": "Ошибка",
    "warning": "Предупреждение",
    "suggestion": "Подсказка",
}

# Excel интерпретирует ведущие = + - @ как формулу: текст из проверяемого
# документа обязан попадать в ячейку как текст, а не исполняться.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_excel_report(issues: list[dict[str, Any]], source_name: str = "") -> bytes:
    """Собирает xlsx и возвращает bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    has_pages = any(issue.get("page_url") for issue in issues)

    headers = ["#", "Страница", "Строка", "Фрагмент", "Тип", "Серьезность", "Описание", "Рекомендация"] if has_pages else ["#", "Строка", "Фрагмент", "Тип", "Серьезность", "Описание", "Рекомендация"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B3A4E", end_color="2B3A4E", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    if has_pages:
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 50
        ws.column_dimensions["H"].width = 30
    else:
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 30

    sorted_issues = sorted(issues, key=lambda x: (x.get("page_url", ""), x.get("line", 0), x.get("column", 0)))

    for row_idx, issue in enumerate(sorted_issues, 2):
        severity = issue.get("severity", "warning")
        fill_color = SEVERITY_COLORS.get(severity, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        category = _issue_category(issue)

        if has_pages:
            values = [
                row_idx - 1,
                issue.get("page_url", ""),
                issue.get("line", 0),
                issue.get("text", ""),
                category,
                SEVERITY_LABELS.get(severity, severity),
                issue.get("message", ""),
                issue.get("replacement", ""),
            ]
            wrap_cols = (4, 7, 8)
        else:
            values = [
                row_idx - 1,
                issue.get("line", 0),
                issue.get("text", ""),
                category,
                SEVERITY_LABELS.get(severity, severity),
                issue.get("message", ""),
                issue.get("replacement", ""),
            ]
            wrap_cols = (3, 6, 7)

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_safe_cell(value))
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx in wrap_cols)

    if source_name:
        ws_info = wb.create_sheet("Информация")
        ws_info["A1"] = "Источник"
        ws_info["B1"] = _safe_cell(source_name)
        ws_info["A2"] = "Всего проблем"
        ws_info["B2"] = len(issues)
        ws_info["A3"] = "Ошибок"
        ws_info["B3"] = sum(1 for i in issues if i.get("severity") == "error")
        ws_info["A4"] = "Предупреждений"
        ws_info["B4"] = sum(1 for i in issues if i.get("severity") == "warning")
        ws_info["A5"] = "Подсказок"
        ws_info["B5"] = sum(1 for i in issues if i.get("severity") == "suggestion")
        for row in ws_info.iter_rows(min_row=1, max_row=5, max_col=2):
            for cell in row:
                cell.font = Font(size=11)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _lt_rule_category(rule: str) -> str:
    rule_lower = rule.lower()
    if "spell" in rule_lower or "typo" in rule_lower:
        return "Орфография"
    if "punct" in rule_lower or "comma" in rule_lower:
        return "Пунктуация"
    return "Грамматика"


def _issue_category(issue: dict[str, Any]) -> str:
    if issue.get("source") == "custom":
        return "Свои правила"
    if issue.get("source") == "style-guide":
        return issue.get("rule_group") or "Style Guide"

    rule = issue.get("rule", "")
    if rule.startswith("LanguageTool."):
        return _lt_rule_category(rule)
    return "Стиль"
